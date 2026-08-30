#!/usr/bin/env python3
"""Extract tissue-domain sensitivity evidence from SRP109982 averaged RPM data."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--expression", type=Path, required=True)
    parser.add_argument("--signature", type=Path, required=True)
    parser.add_argument("--effect-output", type=Path, required=True)
    parser.add_argument("--signature-output", type=Path, required=True)
    parser.add_argument("--summary", type=Path, required=True)
    args = parser.parse_args()

    sheet = openpyxl.load_workbook(args.expression, read_only=True, data_only=True).active
    header = list(next(sheet.iter_rows(min_row=5, max_row=5, values_only=True)))
    header[:3] = ["gene_id", "max_average_rpm", "expressed"]
    rows = sheet.iter_rows(min_row=6, values_only=True)
    expression = pd.DataFrame(rows, columns=header)
    expression = expression[expression.gene_id.astype(str).str.startswith("Solyc")].copy()
    contexts = [
        "Total pericarp", "Septum", "Locular tissue", "Placenta", "Columella", "Seeds",
        "Outer epidermis", "Collenchyma", "Parenchyma", "Vascular tissue", "Inner epidermis",
    ]
    effects = []
    for context in contexts:
        mg_columns = [column for column in expression if isinstance(column, str) and column.startswith(f"{context} MG")]
        ripe_column = f"{context} RR"
        if not mg_columns or ripe_column not in expression:
            raise ValueError(f"Missing MG/RR columns for {context}")
        mg = expression[mg_columns].apply(pd.to_numeric).mean(axis=1)
        ripe = pd.to_numeric(expression[ripe_column])
        effects.append(pd.DataFrame({
            "gene_id": expression.gene_id,
            "context": context,
            "mg_average_rpm": mg,
            "rr_average_rpm": ripe,
            "log2_RR_vs_MG": np.log2(ripe + .1) - np.log2(mg + .1),
            "measured": (mg >= 1) | (ripe >= 1),
            "evidence_resolution": "published_average_of_3_or_4_biological_replicates",
        }))
    long = pd.concat(effects, ignore_index=True)
    signature = pd.read_csv(args.signature)[["gene_id", "effect_source_a"]]
    support = long.merge(signature, on="gene_id", how="inner", validate="many_to_one")
    support["direction_concordant"] = np.sign(support.log2_RR_vs_MG) == np.sign(support.effect_source_a)
    measured = support[support.measured].copy()
    by_gene = measured.groupby("gene_id", as_index=False).agg(
        tissue_context_count=("context", "nunique"),
        tissue_context_direction_consistency=("direction_concordant", "mean"),
        tissue_contexts_concordant=("context", lambda x: ";".join(sorted(set(x[measured.loc[x.index, "direction_concordant"]])))),
        tissue_contexts_measured=("context", lambda x: ";".join(sorted(set(x)))),
    )
    context_rows = []
    for context, group in measured.groupby("context"):
        correlation = group.effect_source_a.rank().corr(group.log2_RR_vs_MG.rank(), method="pearson")
        context_rows.append({
            "context": context,
            "signature_genes_measured": len(group),
            "direction_concordance": float(group.direction_concordant.mean()),
            "spearman_effect_correlation": float(correlation),
        })
    summary = {
        "source": "SRP109982 / PMC5785480 Supplementary Data 3",
        "cultivar": "M82",
        "context_count": len(contexts),
        "contexts": context_rows,
        "admission_status": "tissue_applicability_only",
        "exclusion_from_studyshield": "Only published replicate-averaged RPM values are available in the supplement; individual libraries are not reconstructed here.",
        "claim_boundary": "Supports tissue applicability and heterogeneity mapping, not replicate-level method benchmarking.",
    }
    args.effect_output.parent.mkdir(parents=True, exist_ok=True)
    long.to_csv(args.effect_output, index=False, compression="gzip" if args.effect_output.suffix == ".gz" else None)
    by_gene.to_csv(args.signature_output, index=False)
    args.summary.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
